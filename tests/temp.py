import openpyxl

# Charger le fichier Excel
fichier_excel = "/docs/excel_notes/notes_S1/Traiter_des_donnees.xlsx"
classeur = openpyxl.load_workbook(fichier_excel)

# Sélectionner la feuille de calcul
feuille = classeur.active


# Lire les données
for ligne in feuille.iter_rows(min_row=1, max_row=feuille.max_row, values_only=True):
    # `ligne` = valeurs de chaque cellule dans la ligne
    print(ligne)

# Fermer le classeur après avoir terminé
classeur.close()