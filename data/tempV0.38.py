import os
import openpyxl 

# Charger les fichiers Excel et initialisation des variables
donneesS1 = {}
emplacementcode = os.getcwd()#determine l'emplacement du script dans l'arboressence 
dossierS1 = os.path.dirname(os.path.dirname(emplacementcode))
dossierS1 = os.path.join(dossierS1, "SAE105" , "docs", "excel_notes" , "notes_S1")#permet d'acceder au fichier sur n'importe quelle pc
num_files = 0
for nomfichier in os.listdir(dossierS1):
    if nomfichier.endswith(".xlsx"):#chereche TOUT les fichier en .xlsx du dossier et les charge 
        num_files += 1 #"compte le nombre de fichier .xlsx dans dossierS1"
        fichier_excel = os.path.join(dossierS1, nomfichier)#slelectione les "worksheet active"
        classeur = openpyxl.load_workbook(fichier_excel)# contient plusieur worksheet

        # Sélectionner la feuille de calcul
        feuille = classeur.active

        # Lire et traiter les donnes
        for ligne in feuille.iter_rows(min_row=2, max_row=feuille.max_row, values_only=True):
            # `ligne` = valeurs de chaque cellule dans la ligne
            nom_eleve = ligne[1] + ' ' + ligne[2]
            valeur = ligne[3]
            if nom_eleve in donneesS1:
                donneesS1[nom_eleve] += valeur
            else:
                donneesS1[nom_eleve] = valeur
    
        classeur.close()  # Ferme le classeur

# divise la somme des notes par le nombre de fichier pour obtenir la moyenne
for nom_eleve, somme in donneesS1.items():
    donneesS1[nom_eleve] = round(somme / num_files, 2)#calcule la moyenne et l'arrondie 

emplacementcode = os.getcwd()#determine l'emplacement du script dans l'arboressence 
# Charger les fichiers Excel et initialisation des variables
donneesS2 = {}
dossierS2 = os.path.dirname(os.path.dirname(emplacementcode))
dossierS2 = os.path.join(dossierS2, "SAE105" , "docs", "excel_notes" , "notes_S2")#permet d'acceder au fichier sur n'importe quelle pc
num_files = 0
for nomfichier in os.listdir(dossierS2):
    if nomfichier.endswith(".xlsx"):#chereche TOUT les fichier en .xlsx du dossier et les charge 
        num_files += 1 #"compte le nombre de fichier .xlsx dans dossierS2"
        fichier_excel = os.path.join(dossierS2, nomfichier)#slelectione les "worksheet active"
        classeur = openpyxl.load_workbook(fichier_excel)# contient plusieur worksheet

        # Sélectionner la feuille de calcul
        feuille = classeur.active

        # Lire et traiter les donnes
        for ligne in feuille.iter_rows(min_row=2, max_row=feuille.max_row, values_only=True):
            # `ligne` = valeurs de chaque cellule dans la ligne
            nom_eleve = ligne[1] + ' ' + ligne[2]
            valeur = ligne[3]
            if nom_eleve in donneesS2:
                donneesS2[nom_eleve] += valeur
            else:
                donneesS2[nom_eleve] = valeur
    
        classeur.close()  # Ferme le classeur

# divise la somme des notes par le nombre de fichier pour obtenir la moyenne
for nom_eleve, somme in donneesS2.items():
    donneesS2[nom_eleve] = round(somme / num_files, 2)#calcule la moyenne et l'arrondie 

# Calcule la moyenne des eleve sur les deux semestre
donnees_moyennes = {}
for nom_eleve, valeurS1 in donneesS1.items():
    if nom_eleve in donneesS2:
        valeurS2 = donneesS2[nom_eleve]
        moyenne = round((valeurS1 + valeurS2) / 2, 2)
        donnees_moyennes[nom_eleve] = moyenne

print("S1", donneesS1)
print("S2" , donneesS2)
print("moyenne des deux semestre =", donnees_moyennes)