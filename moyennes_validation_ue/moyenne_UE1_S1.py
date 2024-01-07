import os
import openpyxl 


def calcul_moyennes_notes_UE1_S1():
    # Charger les fichiers Excel et initialisation des variables
    donnees_UE1 = {}
    resultat = {}
    emplacementcode = os.getcwd() #determine l'emplacement du script dans l'arboressence 
    dossier_UE1_S1 = os.path.dirname(os.path.dirname(emplacementcode))
    dossier_UE1_S1 = os.path.join(dossier_UE1_S1, "SAE105" , "docs", "excel_notes" , "notes_S1" , "UE1.1") #permet d'acceder au fichier sur n'importe quelle pc
    num_files = 0
    for nomfichier in os.listdir(dossier_UE1_S1):
        if nomfichier.endswith(".xlsx"): #chereche TOUT les fichier en .xlsx du dossier et les charge 
            num_files += 1 #"compte le nombre de fichier .xlsx dans dossierS1"
            fichier_excel = os.path.join(dossier_UE1_S1, nomfichier) #slelectione les "worksheet active"
            classeur = openpyxl.load_workbook(fichier_excel) # contient plusieur worksheet

            # Sélectionner la feuille de calcul
            feuille = classeur.active

            # Lire et traiter les donnes
            for ligne in feuille.iter_rows(min_row=2, max_row=feuille.max_row, values_only=True):
                # `ligne` = valeurs de chaque cellule dans la ligne
                nom_eleve = ligne[1] + ' ' + ligne[2]
                valeur = ligne[3]
                if nom_eleve in donnees_UE1:
                    donnees_UE1[nom_eleve] += valeur
                else:
                    donnees_UE1[nom_eleve] = valeur
        
            classeur.close()  # Ferme le classeur

    # divise la somme des notes par le nombre de fichier pour obtenir la moyenne
    for nom_eleve, somme in donnees_UE1.items():
        donnees_UE1[nom_eleve] = round(somme / num_files, 2) #calcule la moyenne et l'arrondie 

    # Charger les fichiers Excel et initialisation des variables
    donnees_UE1_S2 = {}
    etatUE1 = {}
    resultat = {}
    emplacementcode = os.getcwd() #determine l'emplacement du script dans l'arboressence 
    dossier_UE1_S2 = os.path.dirname(os.path.dirname(emplacementcode))
    dossier_UE1_S2 = os.path.join(dossier_UE1_S2, "SAE105" , "docs", "excel_notes" , "notes_S2" , "UE1") #permet d'acceder au fichier sur n'importe quelle pc
    num_files = 0
    for nomfichier in os.listdir(dossier_UE1_S2):
        if nomfichier.endswith(".xlsx"): #chereche TOUT les fichier en .xlsx du dossier et les charge 
            num_files += 1 #"compte le nombre de fichier .xlsx dans dossierS2"
            fichier_excel = os.path.join(dossier_UE1_S2, nomfichier) #slelectione les "worksheet active"
            classeur = openpyxl.load_workbook(fichier_excel) # contient plusieur worksheet

            # Sélectionner la feuille de calcul
            feuille = classeur.active

            # Lire et traiter les donnes
            for ligne in feuille.iter_rows(min_row=2, max_row=feuille.max_row, values_only=True):
                # `ligne` = valeurs de chaque cellule dans la ligne
                nom_eleve = ligne[1] + ' ' + ligne[2]
                valeur = ligne[3]
                if nom_eleve in donnees_UE1_S2:
                    donnees_UE1_S2[nom_eleve] += valeur
                else:
                    donnees_UE1_S2[nom_eleve] = valeur
        
            classeur.close()  # Ferme le classeur           

    # divise la somme des notes par le nombre de fichier pour obtenir la moyenne
    for nom_eleve, somme in donnees_UE1_S2.items():
        donnees_UE1_S2[nom_eleve] = round(somme / num_files, 2) #calcule la moyenne et l'arrondie 

    # Calculer la moyenne finale et déterminer si l'UE est valide ou non
    for nom_eleve in donnees_UE1.keys():
        if (donnees_UE1[nom_eleve] + donnees_UE1_S2[nom_eleve]) / 2 >= 10:
            resultat[nom_eleve] = "Valider"
        else:
            resultat[nom_eleve] = "Non valider"

    column_titles = ["Nom", "Prénom", "Moyenne UE1 S1", "Moyenne UE1 S2", "Résultat"]
    data_rows = []
    for nom_eleve in donnees_UE1.keys():
        nom, prenom = nom_eleve.split(' ')
        data_rows.append([nom, prenom, donnees_UE1[nom_eleve], donnees_UE1_S2[nom_eleve], resultat[nom_eleve]])        


    return donnees_UE1, donnees_UE1_S2, resultat, data_rows



def generate_html_file(file_name, title, column_titles, data_rows):
    #Contenu HTML
    html_content = f"""
    <!DOCTYPE html>
    <html lang="en">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>{title}</title>
        <style>


            body {{
                font-family: 'Arial', sans-serif;
            }}
            h1 {{
                color: #0066cc;
            }}
            table {{
                border-collapse: collapse;
                width: 100%;
            }}
            th, td {{
                border: 1px solid #dddddd;
                text-align: left;
                padding: 8px;
            }}
            th {{
                background-color: #f2f2f2;
            }}
        </style>
    </head>
    <body>
        <h1>{title}</h1>
        <table>
            <tr>
                {''.join(f'<th>{col}</th>' for col in column_titles)}
            </tr>
            {''.join(f'<tr>{"".join(f"<td>{data}</td>" for data in row)}</tr>' for row in data_rows)}
        </table>
    </body>
    </html>
    """

    #Ecriture du contenu dans le fichier spécifié
    with open(file_name, "w") as file:
        file.write(html_content)

    print(f"Le fichier {file_name} a été généré avec succès.")